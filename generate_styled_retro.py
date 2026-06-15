import subprocess
import sys

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    install('openpyxl')
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Sprint Retrospective"

# Define styles
title_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
header_fill = PatternFill(start_color="33CCFF", end_color="33CCFF", fill_type="solid")
desc_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")

bold_font = Font(bold=True)
desc_font = Font(italic=True, size=10)

center_align = Alignment(horizontal="center", vertical="center")
wrap_align = Alignment(wrap_text=True, vertical="top")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Row 1: Title
ws.merge_cells('A1:D1')
ws['A1'] = "Sprint Retrospective"
ws['A1'].fill = title_fill
ws['A1'].font = bold_font
ws['A1'].alignment = center_align
ws['A1'].border = thin_border
ws['B1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
ws['C1'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
ws['D1'].border = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Row 2: Headers
headers = ["What went well", "What went poorly", "What ideas do you have", "How should we take action?"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = bold_font
    cell.alignment = center_align
    cell.border = thin_border

# Row 3: Descriptions
descriptions = [
    "This section highlights the successes and positive outcomes from the sprint. It helps the team recognize achievements and identify practices that should be continued.",
    "This section identifies the challenges, roadblocks, or failures encountered during the sprint. It helps pinpoint areas that need improvement or change.",
    "This section is for brainstorming new approaches, tools, or strategies to enhance the team's efficiency, productivity, or project outcomes.",
    "This section outlines specific steps to address the issues and implement the ideas for continuous improvement in future sprints."
]
for col_num, desc in enumerate(descriptions, 1):
    cell = ws.cell(row=3, column=col_num)
    cell.value = desc
    cell.fill = desc_fill
    cell.font = desc_font
    cell.alignment = wrap_align
    cell.border = thin_border

# Row 4: Example data and actual project data
data = [
    [
        "Example : All tasks were completed on time.\nTeam communication was seamless.",
        "Requirements changed mid-sprint.",
        "Plan for a buffer to handle scope changes.",
        "Set stricter deadlines for finalizing requirements"
    ],
    [
        "Model Training: Successfully trained an XGBoost model with data scaling using core voice features (Jitter, Shimmer, PPE).",
        "Voice Recording in Web: Current use of 'sounddevice' for audio recording expects local hardware and won't work easily in cloud deployments.",
        "Refactor Audio Input: Investigate Streamlit-native web audio recorder components (like streamlit-webrtc) to replace sounddevice.",
        "Assign Architecture Team to replace audio recording implementation."
    ],
    [
        "Frontend Implementation: Built a responsive, interactive Streamlit web interface for manual and voice inputs.",
        "Feature Engineering: Currently relying on only 3 features. We have more features in data.csv that could improve model robustness.",
        "Model Evaluation: Conduct a new model training experiment incorporating more features from data.csv.",
        "Data Science Team to run experiment and evaluate accuracy improvement."
    ],
    [
        "Medical Reporting: Implemented automated PDF report generation using reportlab, including risk level and plotting data.",
        "Error Handling: Lack of robust try-except blocks, especially around audio processing and librosa feature extraction.",
        "Add comprehensive error handling around audio workflows.",
        "Backend Developer to implement try-except blocks."
    ]
]

for row_idx, row_data in enumerate(data, start=4):
    for col_idx, cell_value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = cell_value
        cell.alignment = wrap_align
        cell.border = thin_border

# Adjust row heights and column widths
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 65
for row in range(4, 9):
    ws.row_dimensions[row].height = 70

ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 38

wb.save('sprint_retrospective_formatted.xlsx')
print("Successfully generated sprint_retrospective_formatted.xlsx")
